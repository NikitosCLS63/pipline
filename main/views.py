# C:\WebsiteDjSND\main\views.py
from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse
from apps.users.decorators import require_role
from django.utils import timezone
from django.db.models import Sum, Count, F, DecimalField
from apps.products.models import Products, Inventory, Categories
from apps.orders.models import Orders, OrderItems
import csv
import json
from django.db import models
from django.conf import settings
from django.http import StreamingHttpResponse, FileResponse, HttpResponse
import subprocess
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from datetime import datetime
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
import os

def home(request):
    # Получаем товары с изображениями для слайдера
    import json
    slider_products = Products.objects.filter(
        stock_quantity__gt=0,
        status='active'
    ).exclude(
        images__isnull=True
    ).exclude(
        images='[]'
    )[:10]  # Берем первые 10 товаров
    
    slider_images = []
    for product in slider_products:
        # Извлекаем первое изображение из JSONField
        first_image = None
        if product.images:
            try:
                if isinstance(product.images, str):
                    images_data = json.loads(product.images)
                else:
                    images_data = product.images
                
                if isinstance(images_data, list) and len(images_data) > 0:
                    # Берем первое изображение
                    first_image = images_data[0]
                    if isinstance(first_image, str) and not first_image.startswith('blob:'):
                        slider_images.append({
                            'url': first_image,
                            'product_id': product.product_id,
                            'product_name': product.product_name
                        })
            except (json.JSONDecodeError, TypeError, AttributeError):
                pass
        
        # Если не нашли в images, пробуем image_url
        if not first_image and product.image_url:
            slider_images.append({
                'url': product.image_url,
                'product_id': product.product_id,
                'product_name': product.product_name
            })
    
    # Получаем категории для отображения
    categories = Categories.objects.all()[:10]
    
    context = {
        'slider_images': slider_images,
        'categories': categories
    }
    return render(request, 'home.html', context)

def catalog(request):
    return render(request, 'catalog.html')  # Заглушка

def promotions(request):
    return render(request, 'promotions.html')  # Заглушка

def favorites(request):
    """Страница избранного"""
    from apps.cart.models import Wishlists
    from apps.users.models import Customers
    from rest_framework_simplejwt.settings import api_settings
    import logging
    logger = logging.getLogger(__name__)
    
    favorites_list = []
    customer = None
    
    logger.info(f"[FAVORITES] GET /favorites/ - Getting customer")
    logger.info(f"[FAVORITES] request.user: {request.user}")
    logger.info(f"[FAVORITES] All cookies: {list(request.COOKIES.keys())}")
    
    # Сначала пытаемся получить customer_id из JWT токена в Authorization header
    auth_header = request.META.get('HTTP_AUTHORIZATION', '')
    logger.info(f"[FAVORITES] auth_header: {auth_header[:50] if auth_header else 'empty'}")
    
    if auth_header.startswith('Bearer '):
        try:
            from rest_framework_simplejwt.tokens import AccessToken
            token_str = auth_header.split(' ')[1]
            token = AccessToken(token_str)
            customer_id = token.get(api_settings.USER_ID_CLAIM)
            logger.info(f"[FAVORITES] Got customer_id from token: {customer_id} (type: {type(customer_id)})")
            if customer_id:
                customer = Customers.objects.get(customer_id=int(customer_id))
                logger.info(f"[FAVORITES] Found customer from token: {customer.customer_id}")
        except Exception as e:
            logger.error(f"[FAVORITES] Error getting customer from token: {e}")
    
    # Если не получили из токена, пытаемся из cookies (customer_id)
    if not customer:
        customer_id = request.COOKIES.get('customer_id')
        logger.info(f"[FAVORITES] Trying customer_id from cookie: {customer_id} (type: {type(customer_id)})")
        if customer_id:
            try:
                # Конвертируем в int, так как cookie хранится как строка
                customer = Customers.objects.get(customer_id=int(customer_id))
                logger.info(f"[FAVORITES] Found customer from cookie: {customer.customer_id}")
            except (ValueError, TypeError) as e:
                logger.error(f"[FAVORITES] Invalid customer_id type in cookie: {e}")
            except Customers.DoesNotExist:
                logger.error(f"[FAVORITES] Customer with id {customer_id} not found in DB")
            except Exception as e:
                logger.error(f"[FAVORITES] Error getting customer from cookie: {e}")
    
    # Если не получили из cookies, пытаемся по email (для Django sessions)
    if not customer:
        if request.user and hasattr(request.user, 'email') and request.user.email:
            customer = Customers.objects.filter(email=request.user.email).first()
            if customer:
                logger.info(f"[FAVORITES] Found customer from email: {customer.customer_id}")
    
    if customer:
        try:
            wishlists = Wishlists.objects.filter(customer=customer).select_related('product')
            favorites_list = [w.product for w in wishlists]
            logger.info(f"[FAVORITES] SUCCESS! Found {len(favorites_list)} favorites for customer {customer.customer_id}")
        except Exception as e:
            logger.error(f"[FAVORITES] Error getting favorites: {e}")
    else:
        logger.error(f"[FAVORITES] No customer found for request")
    
    context = {'favorites': favorites_list}
    return render(request, 'favorites.html', context)

def orders(request):
    """Страница со всеми заказами пользователя"""
    from apps.users.models import Customers, Users
    from apps.orders.models import Orders
    from apps.users.decorators import get_user_from_request
    from rest_framework_simplejwt.settings import api_settings
    from rest_framework_simplejwt.authentication import JWTAuthentication
    
    orders_list = []
    customer = None
    user = None
    
    # Сначала пытаемся получить через helper функцию
    try:
        user, customer = get_user_from_request(request)
        if customer:
            print(f"[ORDERS] Found customer {customer.customer_id} via get_user_from_request")
    except Exception as e:
        print(f"[ORDERS] Error in get_user_from_request: {e}")
    
    # Если не получили, пытаемся через JWT напрямую
    if not customer:
        try:
            auth = JWTAuthentication()
            header = auth.get_header(request)
            if header:
                raw_token = auth.get_raw_token(header)
                if raw_token:
                    validated_token = auth.get_validated_token(raw_token)
                    customer_id = validated_token.get('customer_id')
                    if customer_id:
                        customer = Customers.objects.get(customer_id=customer_id)
                        print(f"[ORDERS] Found customer {customer.customer_id} via JWT")
        except Exception as e:
            print(f"[ORDERS] Error in JWT auth: {e}")
    
    # Если не получили из JWT, пытаемся из Authorization header напрямую
    if not customer:
        auth_header = request.META.get('HTTP_AUTHORIZATION', '')
        if auth_header.startswith('Bearer '):
            try:
                from rest_framework_simplejwt.tokens import AccessToken
                token_str = auth_header.split(' ')[1]
                token = AccessToken(token_str)
                customer_id = token.get('customer_id')  # Пробуем customer_id напрямую
                if not customer_id:
                    customer_id = token.get(api_settings.USER_ID_CLAIM)
                if customer_id:
                    customer = Customers.objects.get(customer_id=int(customer_id))
                    print(f"[ORDERS] Found customer {customer.customer_id} via AccessToken")
            except Exception as e:
                print(f"[ORDERS] Error getting customer from token: {e}")
    
    # Если не получили из токена, пытаемся по email (для Django sessions)
    if not customer:
        if request.user and hasattr(request.user, 'email') and request.user.email:
            customer = Customers.objects.filter(email=request.user.email).first()
            if customer:
                print(f"[ORDERS] Found customer {customer.customer_id} via email")
    
    # Если всё равно нет, пытаемся получить из cookies
    if not customer:
        customer_id = request.COOKIES.get('customer_id')
        if customer_id:
            try:
                customer = Customers.objects.get(customer_id=int(customer_id))
                print(f"[ORDERS] Found customer {customer.customer_id} via cookie")
            except Exception as e:
                print(f"[ORDERS] Error getting customer from cookie: {e}")
    
    if customer:
        try:
            # Используем prefetch_related для загрузки orderitems и products
            from apps.orders.models import OrderItems
            orders_list = Orders.objects.filter(
                customer=customer
            ).select_related(
                'shipping_address', 'customer'
            ).prefetch_related(
                'orderitems_set__product'
            ).order_by('-order_date')
            
            print(f"[ORDERS] Found {len(orders_list)} orders for customer {customer.customer_id}")
            # Логируем ID заказов для отладки
            order_ids = [o.order_id for o in orders_list]
            print(f"[ORDERS] Order IDs: {order_ids}")
            
            # Проверяем, что заказы действительно загружены
            for order in orders_list:
                items_count = order.orderitems_set.count()
                print(f"[ORDERS] Order {order.order_id}: {items_count} items, total={order.total_amount}")
        except Exception as e:
            print(f"[ORDERS] Error getting orders: {e}")
            import traceback
            traceback.print_exc()
            orders_list = []
    else:
        print(f"[ORDERS] No customer found for request")
        print(f"[ORDERS] DEBUG: Request headers: {dict(request.META)}")
        print(f"[ORDERS] DEBUG: Cookies: {dict(request.COOKIES)}")
        print(f"[ORDERS] DEBUG: All orders count: {Orders.objects.count()}")
        orders_list = []
    
    context = {'orders': orders_list, 'customer_id': customer.customer_id if customer else None}
    return render(request, 'orders.html', context)

def register(request):
    return render(request, 'register.html')  # Заглушка

def login(request):
    return render(request, 'login.html')  # Заглушка


# Админ-панель - доступна только для admin и employee
@require_role('admin', 'employee')
def admin_panel(request):
    from apps.users.decorators import get_user_from_request
    user, customer = get_user_from_request(request)
    user_role = user.role.role_name if (user and user.role) else 'client'
    is_admin = user_role == 'admin'
    is_employee = user_role == 'employee'
    
    return render(request, 'admin/admin_panel.html', {
        'user_role': user_role,
        'is_admin': is_admin,
        'is_employee': is_employee
    })


# Управление пользователями - только для admin
@require_role('admin')
def admin_users(request):
    return render(request, 'admin/users.html')


# Редактирование пользователя - только для admin
@require_role('admin')
def admin_user_edit(request, user_id):
    from apps.users.models import Users, Customers, Roles
    from apps.users.decorators import get_user_from_request
    
    try:
        user = Users.objects.select_related('customer', 'role').get(users_id=user_id)
        customer = user.customer
        roles = Roles.objects.all()
    except Users.DoesNotExist:
        from django.http import Http404
        raise Http404("Пользователь не найден")
    
    context = {
        'user': user,
        'customer': customer,
        'roles': roles,
        'user_id': user_id
    }
    
    return render(request, 'admin/user_edit.html', context)


# Админ: скачать бэкап базы данных (только для админа)
@require_role('admin')
def admin_backup_db(request):
    db = settings.DATABASES.get('default', {})
    engine = db.get('ENGINE', '')

    # SQLite: просто отдать файл
    if 'sqlite3' in engine:
        db_path = db.get('NAME')
        if not db_path or not os.path.exists(db_path):
            return HttpResponse('Database file not found', status=404)
        return FileResponse(open(db_path, 'rb'), as_attachment=True, filename=os.path.basename(db_path))

    # PostgreSQL: попытаться вызвать pg_dump и стримить вывод
    if 'postgresql' in engine or 'postgres' in engine:
        host = db.get('HOST', 'localhost')
        port = str(db.get('PORT', '5432'))
        name = db.get('NAME')
        user = db.get('USER')
        password = db.get('PASSWORD', '')

        pg_dump_cmd = ['pg_dump', '--host', host, '--port', port, '--username', user, '--format', 'custom', name]
        env = os.environ.copy()
        if password:
            env['PGPASSWORD'] = password
        try:
            proc = subprocess.Popen(pg_dump_cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, env=env)
        except FileNotFoundError:
            return HttpResponse('pg_dump not found on server. Install PostgreSQL client utilities.', status=500)

        response = StreamingHttpResponse(proc.stdout, content_type='application/octet-stream')
        response['Content-Disposition'] = f'attachment; filename="{name}_backup.dump"'
        return response

    return HttpResponse('Unsupported database engine for backup', status=500)


# Управление запасами (добавление записей, увеличение количества)
@require_role('admin', 'employee')
def admin_inventory(request):
    if request.method == 'POST':
        try:
            product_id = int(request.POST.get('product_id'))
            qty = int(request.POST.get('quantity'))
            warehouse = request.POST.get('warehouse_id') or None
        except Exception:
            return JsonResponse({'error': 'Invalid input'}, status=400)

        try:
            product = Products.objects.get(product_id=product_id)
        except Products.DoesNotExist:
            return JsonResponse({'error': 'Product not found'}, status=404)

        # Создаём запись в Inventory
        inv = Inventory()
        inv.product_id = product.product_id
        inv.quantity = qty
        if warehouse:
            try:
                inv.warehouse_id = int(warehouse)
            except Exception:
                inv.warehouse_id = None
        inv.updated_at = timezone.now()
        try:
            inv.save()
        except Exception as e:
            print(f"[INVENTORY] Ошибка при сохранении инвентаря: {str(e)}")
            pass

        # Обновляем stock_quantity у продукта атомарно используя F()
        try:
            from django.db.models import F
            if product.stock_quantity is None or product.stock_quantity == 0:
                # Если NULL или 0, сначала установим значение
                product.stock_quantity = qty
                product.save(update_fields=['stock_quantity'])
                print(f"[STOCK_UPDATE] Товар {product.product_id}: установлено stock_quantity={qty}")
            else:
                # Атомарное увеличение
                Products.objects.filter(product_id=product.product_id).update(
                    stock_quantity=F('stock_quantity') + qty
                )
                product.refresh_from_db()
                print(f"[STOCK_UPDATE] Товар {product.product_id}: увеличено stock_quantity на {qty}, новое значение={product.stock_quantity}")
        except Exception as e:
            print(f"[STOCK_ERROR] Ошибка при обновлении stock_quantity для товара {product.product_id}: {str(e)}")
            pass

        return redirect('admin_panel')

    # GET: показать форму
    products = Products.objects.all().order_by('product_name')[:1000]
    # select_related to ensure product relation is available for template
    try:
        recent_inv = Inventory.objects.select_related('product').all().order_by('-updated_at')[:50]
    except Exception:
        recent_inv = Inventory.objects.all().order_by('-updated_at')[:50]
    return render(request, 'admin/inventory.html', {'products': products, 'recent_inv': recent_inv})


# Управление заказами (список и обновление статуса)
# Управление заказами (список и обновление статуса)
@require_role('admin', 'employee')
def admin_orders_manage(request):
    message = None
    message_type = None
    
    if request.method == 'POST':
        order_id = request.POST.get('order_id')
        new_status = request.POST.get('status')
        tracking = request.POST.get('tracking_number') or None
        
        if not order_id or not new_status:
            message = 'Ошибка: не указан ID заказа или статус'
            message_type = 'error'
        else:
            try:
                order = Orders.objects.get(order_id=order_id)
                order.status = new_status

                if tracking:
                    order.tracking_number = tracking
                
                # Используем update_fields для более точного обновления
                update_fields = ['status']
                if tracking:
                    update_fields.append('tracking_number')
                
                order.save(update_fields=update_fields)
                message = f'Статус заказа #{order_id} успешно обновлен'
                message_type = 'success'

            except Orders.DoesNotExist:
                message = f'Заказ #{order_id} не найден'
                message_type = 'error'

            except Exception as e:
                message = f'Ошибка при обновлении заказа: {str(e)}'
                message_type = 'error'

    # Получаем все заказы с предварительной загрузкой связанных данных
    orders = Orders.objects.select_related(
        'customer', 
        'shipping_address'
    ).prefetch_related(
        'orderitems_set'
    ).order_by('-order_date')[:200]

    return render(request, 'admin/orders_manage.html', {
        'orders': orders,
        'message': message,
        'message_type': message_type
    })


# Analytics page for admin: revenue, orders count, items sold
@require_role('admin')
def admin_analytics(request):
    now = timezone.now()

    def series_for_days(days):
        labels = []
        revenue = []
        orders_count = []
        items_sold = []
        for i in range(days-1, -1, -1):
            day = (now - timezone.timedelta(days=i)).date()
            labels.append(day.strftime('%Y-%m-%d'))
            start = timezone.datetime.combine(day, timezone.datetime.min.time(), tzinfo=timezone.utc)
            end = timezone.datetime.combine(day, timezone.datetime.max.time(), tzinfo=timezone.utc)
            # Revenue
            rev = Orders.objects.filter(order_date__range=(start, end)).aggregate(total=Sum('total_amount'))['total'] or 0
            revenue.append(float(rev))
            # Orders count
            cnt = Orders.objects.filter(order_date__range=(start, end)).aggregate(count=Count('order_id'))['count'] or 0
            orders_count.append(int(cnt))
            # Items sold
            items = OrderItems.objects.filter(order__order_date__range=(start, end)).aggregate(total=Sum('quantity'))['total'] or 0
            items_sold.append(int(items))

        return {'labels': labels, 'revenue': revenue, 'orders': orders_count, 'items': items_sold}

    week = series_for_days(7)
    month = series_for_days(30)
    year = series_for_days(365)

    # Передаем Python объекты, а не JSON строки, так как json_script сам сериализует
    context = {
        'week': week,
        'month': month,
        'year': year,
    }
    return render(request, 'admin/analytics.html', context)


# CSV Export Orders
@require_role('admin')
def admin_export_orders_csv(request):
    orders = Orders.objects.all().order_by('-order_date')
    
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="orders_export.csv"'
    writer = csv.writer(response)
    writer.writerow(['Order ID', 'Customer', 'Date', 'Total Amount', 'Status', 'Payment Status'])
    
    for order in orders:
        customer_name = getattr(order.customer, 'customer_name', f'Customer {order.customer_id}')
        writer.writerow([
            order.order_id,
            customer_name,
            order.order_date.strftime('%Y-%m-%d %H:%M:%S'),
            f'{order.total_amount:.2f}',
            order.status,
            order.payment_status or 'N/A'
        ])
    
    return response


# CSV Export Reports (export sales data)
@require_role('admin')
def admin_import_reports(request):
    """
    Экспорт отчётов в CSV формате.
    Экспортирует данные о проданных товарах с колонками: product_id, quantity, date
    """
    # Если есть параметр download, экспортируем CSV
    if request.GET.get('download') == '1':
        # Получаем параметры фильтрации
        period = request.GET.get('period', 'all')  # all, week, month, year
        
        # Определяем период для фильтрации
        now = timezone.now()
        if period == 'week':
            start_date = (now - timezone.timedelta(days=7)).date()
        elif period == 'month':
            start_date = (now - timezone.timedelta(days=30)).date()
        elif period == 'year':
            start_date = (now - timezone.timedelta(days=365)).date()
        else:
            start_date = None
        
        # Получаем данные о проданных товарах с дополнительной информацией
        if start_date:
            order_items = OrderItems.objects.filter(
                order__order_date__date__gte=start_date
            ).select_related('product', 'product__category', 'order', 'order__customer').order_by('-order__order_date')
        else:
            order_items = OrderItems.objects.select_related('product', 'product__category', 'order', 'order__customer').order_by('-order__order_date')
            
        # Создаём CSV ответ с BOM для правильного отображения UTF-8 в Excel
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        filename = f'reports_export_{period}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Создаём CSV writer
        writer = csv.writer(response, delimiter=',', quoting=csv.QUOTE_MINIMAL)
        
        # Записываем заголовки с подробной информацией
        writer.writerow([
            'ID заказа',
            'ID товара',
            'Название товара',
            'Категория',
            'Количество',
            'Цена за единицу (₽)',
            'Сумма (₽)',
            'Дата заказа',
            'Статус заказа',
            'Способ оплаты',
            'Статус платежа',
            'Клиент (email)'
        ])
        
        # Записываем данные
        for item in order_items:
            product_name = item.product.product_name if item.product else 'Неизвестно'
            category_name = item.product.category.category_name if (item.product and item.product.category) else 'Без категории'
            order_date = item.order.order_date
            # Используем формат даты, который Excel правильно откроет
            date_str = order_date.strftime('%d.%m.%Y %H:%M')
            total_sum = float(item.quantity) * float(item.price_at_purchase)
            
            # Перевод статусов на русский
            status_map = {
                'new': 'Новый',
                'processing': 'В обработке',
                'shipped': 'Отправлен',
                'delivered': 'Доставлен',
                'cancelled': 'Отменен',
                'pending': 'Ожидает оплаты'
            }
            order_status = status_map.get(item.order.status, item.order.status)
            
            payment_method_map = {
                'card': 'Банковская карта',
                'cash': 'Наличные',
                'online': 'Онлайн оплата',
                'on_delivery': 'При получении'
            }
            payment_method = payment_method_map.get(item.order.payment_method, item.order.payment_method or 'Не указан')
            
            payment_status_map = {
                'completed': 'Оплачено',
                'pending': 'Ожидает оплаты',
                'failed': 'Ошибка оплаты',
                'refunded': 'Возврат'
            }
            payment_status = payment_status_map.get(item.order.payment_status, item.order.payment_status or 'Не указан')
            
            customer_email = item.order.customer.email if item.order.customer else f'Клиент #{item.order.customer_id}'
            
            writer.writerow([
                item.order.order_id,
                item.product_id,
                product_name,
                category_name,
                item.quantity,
                f'{item.price_at_purchase:.2f}',
                f'{total_sum:.2f}',
                date_str,
                order_status,
                payment_method,
                payment_status,
                customer_email
            ])
        
        return response
    
    # Иначе показываем страницу с формой
    return render(request, 'admin/import_reports.html')


# Export Products to CSV (all or by category)
@require_role('admin')
def admin_export_products_csv(request):
    category_id = request.GET.get('category_id')
    
    if category_id:
        try:
            products = Products.objects.filter(category_id=category_id).order_by('product_name')
        except Exception:
            products = Products.objects.all().order_by('product_name')
    else:
        products = Products.objects.all().order_by('product_name')
    
    # Always download CSV when there's a category_id or download=1
    if category_id or request.GET.get('download') == '1':
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="products_export.csv"'
        writer = csv.writer(response)
        writer.writerow(['Product ID', 'Product Name', 'Price', 'Category', 'Brand', 'Stock Quantity', 'Status', 'SKU'])
        
        for product in products:
            category_name = product.category.category_name if product.category else 'N/A'
            brand_name = product.brand.brand_name if product.brand else 'N/A'
            writer.writerow([
                product.product_id,
                product.product_name,
                f'{product.price:.2f}',
                category_name,
                brand_name,
                product.stock_quantity or 0,
                product.status or 'N/A',
                product.sku
            ])
        
        return response
    
    # Show export page with options (only if no download requested)
    categories = Categories.objects.all().order_by('category_name')
    return render(request, 'admin/export_products.html', {'categories': categories})


# Review Moderation: view, approve, reject, delete reviews
@require_role('admin', 'employee')
def admin_moderate_reviews(request):
    from apps.reviews.models import Reviews
    
    # Handle approve/reject/delete actions
    if request.method == 'POST':
        action = request.POST.get('action')
        review_id = request.POST.get('review_id')
        
        try:
            review = Reviews.objects.get(review_id=review_id)
            
            if action == 'approve':
                review.status = 'approved'
                review.save()
            elif action == 'reject':
                review.status = 'rejected'
                review.save()
            elif action == 'delete':
                review.delete()
        except Reviews.DoesNotExist:
            pass
        
        return redirect('admin_moderate_reviews')
    
    # Get all reviews with related data
    try:
        reviews = Reviews.objects.select_related('product', 'customer').all().order_by('-publication_date')
    except Exception:
        reviews = Reviews.objects.all().order_by('-publication_date')
    
    # Filter by product if requested
    product_id = request.GET.get('product_id')
    if product_id:
        try:
            reviews = reviews.filter(product_id=int(product_id))
        except Exception:
            pass
    
    # Filter by approval status if requested
    status_filter = request.GET.get('status')
    if status_filter == 'pending':
        reviews = reviews.exclude(status='approved').exclude(status='rejected')
    elif status_filter == 'approved':
        reviews = reviews.filter(status='approved')
    elif status_filter == 'rejected':
        reviews = reviews.filter(status='rejected')
    
    context = {
        'reviews': reviews[:500],  # Limit to 500 for performance
        'product_id': product_id,
        'status': status_filter,
    }
    return render(request, 'admin/moderate_reviews.html', context)


@require_role('admin')
def admin_audit_log(request):
    """Просмотр логов аудита для администраторов"""
    from apps.analytics.models import AuditLog
    from django.core.paginator import Paginator
    from django.db.models import Q

    # Получаем параметры фильтрации
    action_filter = request.GET.get('action', '')
    table_filter = request.GET.get('table', '')
    user_filter = request.GET.get('user', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    # Базовый запрос
    audit_logs = AuditLog.objects.select_related('user').order_by('-timestamp')

    # Применяем фильтры
    if action_filter:
        audit_logs = audit_logs.filter(action_type=action_filter)

    if table_filter:
        audit_logs = audit_logs.filter(table_name=table_filter)

    if user_filter:
        audit_logs = audit_logs.filter(
            Q(user__first_name__icontains=user_filter) |
            Q(user__last_name__icontains=user_filter) |
            Q(user__email__icontains=user_filter)
        )

    if date_from:
        audit_logs = audit_logs.filter(timestamp__date__gte=date_from)

    if date_to:
        audit_logs = audit_logs.filter(timestamp__date__lte=date_to)

    # Пагинация
    paginator = Paginator(audit_logs, 50)  # 50 записей на страницу
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Получаем уникальные значения для фильтров
    actions = AuditLog.objects.values_list('action_type', flat=True).distinct()
    tables = AuditLog.objects.values_list('table_name', flat=True).distinct()

    context = {
        'audit_logs': page_obj,
        'actions': actions,
        'tables': tables,
        'action_filter': action_filter,
        'table_filter': table_filter,
        'user_filter': user_filter,
        'date_from': date_from,
        'date_to': date_to,
    }

    return render(request, 'admin/audit.html', context)


# Sales Report: products sold, revenue by product
@require_role('admin')
def admin_sales_report(request):
    period = request.GET.get('period', 'month')  # week, month, year
    
    # Calculate date range
    now = timezone.now()
    if period == 'week':
        start_date = (now - timezone.timedelta(days=7)).date()
    elif period == 'year':
        start_date = (now - timezone.timedelta(days=365)).date()
    else:  # month
        start_date = (now - timezone.timedelta(days=30)).date()
    
    # Query: aggregate sales by product
    from django.db.models import F
    sales_data = OrderItems.objects.filter(
        order__order_date__date__gte=start_date
    ).values('product_id', 'product__product_name', 'product__category__category_name').annotate(
        total_quantity=Count('item_id'),
        total_revenue=Sum(F('quantity') * F('price_at_purchase'), output_field=models.DecimalField())
    ).order_by('-total_revenue')
    
    # Рассчитываем общую выручку и количество
    total_revenue = sum(float(item['total_revenue'] or 0) for item in sales_data)
    total_quantity = sum(int(item['total_quantity'] or 0) for item in sales_data)
    
    context = {
        'sales_data': sales_data,
        'period': period,
        'start_date': start_date,
        'now': now.date(),
        'total_revenue': total_revenue,
        'total_quantity': total_quantity,
    }
    return render(request, 'admin/sales_report.html', context)


# Export Sales Report to CSV
@require_role('admin')
def admin_export_sales_report_csv(request):
    period = request.GET.get('period', 'month')
    
    now = timezone.now()
    if period == 'week':
        start_date = (now - timezone.timedelta(days=7)).date()
    elif period == 'year':
        start_date = (now - timezone.timedelta(days=365)).date()
    else:
        start_date = (now - timezone.timedelta(days=30)).date()
    
    sales_data = OrderItems.objects.filter(
        order__order_date__date__gte=start_date
    ).values('product_id', 'product__product_name', 'product__category__category_name').annotate(
        total_quantity=Count('item_id'),
        total_revenue=Sum(F('quantity') * F('price_at_purchase'), output_field=models.DecimalField())
    ).order_by('-total_revenue')
    
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="sales_report_{period}.csv"'
    writer = csv.writer(response)
    writer.writerow(['Product ID', 'Product Name', 'Category', 'Quantity Sold', 'Total Revenue'])
    
    total_revenue_sum = 0
    for item in sales_data:
        revenue = item['total_revenue'] or 0
        total_revenue_sum += revenue
        writer.writerow([
            item['product_id'],
            item['product__product_name'],
            item['product__category__category_name'] or 'N/A',
            item['total_quantity'],
            f'{revenue:.2f}'
        ])
    
    # Add summary row
    writer.writerow([])
    writer.writerow(['ИТОГО', '', '', '', f'{total_revenue_sum:.2f}'])
    
    return response


# Экспорт отчёта по продажам в Excel
@require_role('admin')
def admin_export_sales_report_excel(request):
    period = request.GET.get('period', 'month')
    
    # Используем московское время (уже настроено в settings.py как TIME_ZONE = 'Europe/Moscow')
    # Для корректного отображения времени используем zoneinfo из стандартной библиотеки
    try:
        from zoneinfo import ZoneInfo
        moscow_tz = ZoneInfo('Europe/Moscow')
        now = timezone.now().astimezone(moscow_tz)
    except ImportError:
        # Если zoneinfo не доступен (Python < 3.9), используем настройки Django
        # timezone.now() уже учитывает TIME_ZONE из settings.py
        now = timezone.now()
    
    if period == 'week':
        start_date = (now - timezone.timedelta(days=7)).date()
        period_name = 'неделя'
    elif period == 'year':
        start_date = (now - timezone.timedelta(days=365)).date()
        period_name = 'год'
    else:
        start_date = (now - timezone.timedelta(days=30)).date()
        period_name = 'месяц'
    
    sales_data = OrderItems.objects.filter(
        order__order_date__date__gte=start_date
    ).values('product_id', 'product__product_name', 'product__category__category_name').annotate(
        total_quantity=Count('item_id'),
        total_revenue=Sum(F('quantity') * F('price_at_purchase'), output_field=models.DecimalField())
    ).order_by('-total_revenue')
    
    # Создаём workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Продажи'
    
    # Стили
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    total_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    total_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Заголовок
    ws['A1'] = 'ОТЧЕТ ПО ПРОДАЖАМ'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws['A2'] = f'Период: последний {period_name}'
    ws['A2'].font = Font(size=10, italic=True)
    ws.merge_cells('A2:E2')
    
    ws['A3'] = f'Дата: {now.strftime("%d.%m.%Y %H:%M")}'
    ws['A3'].font = Font(size=10, italic=True)
    ws.merge_cells('A3:E3')
    
    # Пустая строка
    ws.append([])
    
    # Заголовки таблицы
    headers = ['ID Товара', 'Название Товара', 'Категория', 'Кол-во Продано', 'Выручка']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Данные
    row = 6
    total_revenue_sum = 0
    
    for item in sales_data:
        revenue = item['total_revenue'] or 0
        total_revenue_sum += revenue
        
        ws.cell(row=row, column=1).value = item['product_id']
        ws.cell(row=row, column=2).value = item['product__product_name']
        ws.cell(row=row, column=3).value = item['product__category__category_name'] or 'N/A'
        ws.cell(row=row, column=4).value = item['total_quantity']
        ws.cell(row=row, column=5).value = revenue
        
        # Стили для данных
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            if col == 4:  # Кол-во
                cell.alignment = Alignment(horizontal='right')
            elif col == 5:  # Выручка
                cell.alignment = Alignment(horizontal='right')
                cell.number_format = '#,##0.00'
            else:
                cell.alignment = Alignment(horizontal='left')
        
        row += 1
    
    # Строка итога
    ws.cell(row=row, column=1).value = 'ИТОГО'
    ws.cell(row=row, column=1).font = total_font
    ws.cell(row=row, column=1).fill = total_fill
    ws.cell(row=row, column=5).value = total_revenue_sum
    ws.cell(row=row, column=5).font = total_font
    ws.cell(row=row, column=5).fill = total_fill
    ws.cell(row=row, column=5).number_format = '#,##0.00'
    
    for col in range(1, 6):
        cell = ws.cell(row=row, column=col)
        cell.border = border
        cell.fill = total_fill
    
    # Ширина колонок
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    # Сохраняем в BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="sales_report_{period}_{now.strftime("%d_%m_%Y")}.xlsx"'
    
    return response


# Экспорт отчёта по продажам в PDF
@require_role('admin')
def admin_export_sales_report_pdf(request):
    period = request.GET.get('period', 'month')
    
    # Используем московское время (уже настроено в settings.py как TIME_ZONE = 'Europe/Moscow')
    # Для корректного отображения времени используем zoneinfo из стандартной библиотеки
    try:
        from zoneinfo import ZoneInfo
        moscow_tz = ZoneInfo('Europe/Moscow')
        now = timezone.now().astimezone(moscow_tz)
    except ImportError:
        # Если zoneinfo не доступен (Python < 3.9), используем настройки Django
        # timezone.now() уже учитывает TIME_ZONE из settings.py
        now = timezone.now()
    if period == 'week':
        start_date = (now - timezone.timedelta(days=7)).date()
        period_name = 'неделя'
    elif period == 'year':
        start_date = (now - timezone.timedelta(days=365)).date()
        period_name = 'год'
    else:
        start_date = (now - timezone.timedelta(days=30)).date()
        period_name = 'месяц'
    
    sales_data = list(OrderItems.objects.filter(
        order__order_date__date__gte=start_date
    ).values('product_id', 'product__product_name', 'product__category__category_name').annotate(
        total_quantity=Count('item_id'),
        total_revenue=Sum(F('quantity') * F('price_at_purchase'), output_field=models.DecimalField())
    ).order_by('-total_revenue'))
    
    # Создаём PDF документ
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=0.5*inch,
        leftMargin=0.5*inch,
        topMargin=0.75*inch,
        bottomMargin=0.75*inch,
        title=f'Отчет по продажам {period_name}'
    )
    
    # Регистрируем шрифт с поддержкой кириллицы
    # Используем системные шрифты Windows с поддержкой кириллицы
    cyrillic_font = 'Helvetica'  # По умолчанию
    
    # Пути к системным шрифтам Windows с поддержкой кириллицы
    font_paths = [
        r'C:\Windows\Fonts\arial.ttf',  # Arial - поддерживает кириллицу
        r'C:\Windows\Fonts\arialuni.ttf',  # Arial Unicode MS - полная поддержка кириллицы
        r'C:\Windows\Fonts\times.ttf',  # Times New Roman - поддерживает кириллицу
        r'C:\Windows\Fonts\tahoma.ttf',  # Tahoma - поддерживает кириллицу
        r'C:\Windows\Fonts\calibri.ttf',  # Calibri - поддерживает кириллицу
    ]
    
    font_loaded = False
    for font_path in font_paths:
        if os.path.exists(font_path):
            try:
                # Определяем имя шрифта
                font_base = os.path.basename(font_path).lower()
                if 'arial' in font_base and 'uni' in font_base:
                    font_name = 'ArialUnicode'
                elif 'arial' in font_base:
                    font_name = 'ArialCyrillic'
                elif 'times' in font_base:
                    font_name = 'TimesCyrillic'
                elif 'tahoma' in font_base:
                    font_name = 'TahomaCyrillic'
                elif 'calibri' in font_base:
                    font_name = 'CalibriCyrillic'
                else:
                    font_name = 'CyrillicFont'
                
                # Регистрируем TTF шрифт с поддержкой кириллицы
                pdfmetrics.registerFont(TTFont(font_name, font_path))
                cyrillic_font = font_name
                font_loaded = True
                print(f"Загружен шрифт с поддержкой кириллицы: {font_path} как {font_name}")
                break
            except Exception as e:
                print(f"Не удалось загрузить шрифт {font_path}: {e}")
                continue
    
    # Если системные шрифты не найдены, пробуем UnicodeCIDFont
    if not font_loaded:
        try:
            pdfmetrics.registerFont(UnicodeCIDFont('HeiseiMin-W3'))
            cyrillic_font = 'HeiseiMin-W3'
            font_loaded = True
            print("Загружен шрифт HeiseiMin-W3 для поддержки кириллицы")
        except:
            try:
                pdfmetrics.registerFont(UnicodeCIDFont('HeiseiKakuGo-W5'))
                cyrillic_font = 'HeiseiKakuGo-W5'
                font_loaded = True
                print("Загружен шрифт HeiseiKakuGo-W5 для поддержки кириллицы")
            except:
                print("Не удалось загрузить шрифт с поддержкой кириллицы, будет использоваться стандартный шрифт")
                cyrillic_font = 'Helvetica'
    
    # Стили с темным цветом текста
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1a1a1a'),  # Темный цвет вместо синего
        spaceAfter=30,
        alignment=TA_CENTER,
        fontName=cyrillic_font
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#333333'),  # Темно-серый вместо светло-серого
        spaceAfter=12,
        alignment=TA_CENTER,
        fontName=cyrillic_font
    )
    
    # Элементы документа
    elements = []
    
    # Заголовок - всегда используем кириллицу
    title_text = 'ОТЧЕТ ПО ПРОДАЖАМ'
    period_text = f'Период: последний {period_name} | Дата: {now.strftime("%d.%m.%Y %H:%M")}'
    
    # Для кириллицы используем Paragraph с правильной кодировкой
    elements.append(Paragraph(title_text, title_style))
    elements.append(Paragraph(period_text, subtitle_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Создаем стили для текста в таблице (для правильной обработки кириллицы)
    table_normal_style = ParagraphStyle(
        'TableNormal',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#1a1a1a'),
        fontName=cyrillic_font,
        leading=11
    )
    
    table_header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.whitesmoke,
        fontName=cyrillic_font,
        alignment=TA_CENTER
    )
    
    table_summary_style = ParagraphStyle(
        'TableSummary',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#1a1a1a'),
        fontName=cyrillic_font,
        alignment=TA_RIGHT
    )
    
    # Подготовка данных таблицы - используем Paragraph для правильной обработки кириллицы
    table_data = [
        [
            Paragraph('ID', table_header_style),
            Paragraph('Название Товара', table_header_style),
            Paragraph('Категория', table_header_style),
            Paragraph('Кол-во', table_header_style),
            Paragraph('Выручка (RUB)', table_header_style)
        ]
    ]
    
    total_revenue_sum = 0
    total_quantity_sum = 0
    for item in sales_data:
        revenue = float(item['total_revenue'] or 0)
        quantity = int(item['total_quantity'] or 0)
        total_revenue_sum += revenue
        total_quantity_sum += quantity
        
        # Получаем названия напрямую из БД (данные уже в UTF-8)
        product_name_raw = item.get('product__product_name') or 'N/A'
        category_name_raw = item.get('product__category__category_name') or 'N/A'
        
        # Преобразуем в строку и обрезаем - данные уже в правильной кодировке UTF-8
        product_name = str(product_name_raw)[:40]
        category_name = str(category_name_raw)[:20]
        
        # Используем Paragraph для каждого текстового элемента, чтобы ReportLab правильно обработал UTF-8
        table_data.append([
            str(item['product_id']),
            Paragraph(product_name, table_normal_style),
            Paragraph(category_name, table_normal_style),
            str(quantity),
            Paragraph(f'{revenue:.2f}', table_normal_style)
        ])
    
    # Строка итога - тоже используем Paragraph для кириллицы
    table_data.append([
        '',
        '',
        '',
        Paragraph('ИТОГО:', table_summary_style),
        Paragraph(f'{total_revenue_sum:.2f}', table_summary_style)
    ])
    
    # Создаём таблицу
    table = Table(table_data, colWidths=[0.8*inch, 2.5*inch, 1.2*inch, 0.8*inch, 1.2*inch])
    
    # Стили таблицы
    table_style = TableStyle([
        # Заголовок
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), cyrillic_font),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        
        # Строка итога
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#D9E1F2')),
        ('FONTNAME', (0, -1), (-1, -1), cyrillic_font),
        ('FONTSIZE', (0, -1), (-1, -1), 10),
        ('TOPPADDING', (0, -1), (-1, -1), 10),
        ('BOTTOMPADDING', (0, -1), (-1, -1), 10),
        
        # Основные ячейки с темным цветом текста
        ('ALIGN', (0, 1), (-1, -2), 'LEFT'),
        ('ALIGN', (3, 1), (-1, -2), 'RIGHT'),  # Выравнивание вправо для чисел
        ('ALIGN', (0, -1), (-1, -1), 'LEFT'),
        ('ALIGN', (3, -1), (-1, -1), 'RIGHT'),  # Выравнивание вправо для итога
        
        ('TEXTCOLOR', (0, 1), (-1, -2), colors.HexColor('#1a1a1a')),  # Темный цвет для данных
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.HexColor('#1a1a1a')),  # Темный цвет для итога
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        
        # Линии
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('LINEABOVE', (0, -1), (-1, -1), 2, colors.HexColor('#4472C4')),
    ])
    
    table.setStyle(table_style)
    elements.append(table)
    
    # Дополнительная информация
    elements.append(Spacer(1, 0.2*inch))
    
    summary_style = ParagraphStyle(
        'Summary',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#1a1a1a'),  # Темный цвет вместо серого
        alignment=TA_LEFT,
        spaceAfter=6,
        fontName=cyrillic_font
    )
    
    # Всегда используем кириллицу в итогах
    elements.append(Paragraph(
        f'<b>Всего товаров продано:</b> {total_quantity_sum} шт.',
        summary_style
    ))
    elements.append(Paragraph(
        f'<b>Общая выручка:</b> {total_revenue_sum:.2f} RUB',
        summary_style
    ))
    elements.append(Paragraph(
        f'<b>Период анализа:</b> с {start_date.strftime("%d.%m.%Y")} по {now.date().strftime("%d.%m.%Y")}',
        summary_style
    ))
    
    # Генерируем PDF
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="sales_report_{period}_{now.strftime("%d_%m_%Y")}.pdf"'
    
    return response